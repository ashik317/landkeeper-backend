import io
from datetime import date

from django.contrib.auth.hashers import make_password
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.exceptions import NotFound
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, ListAPIView
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from apps.organisation.models import OrganisationUser
from apps.organisation.enums import OrganisationRoleChoices
from apps.property.models import (
    Property,
    Mortgage,
    Tenant,
    ComplianceAndCertification,
    UploadDocument,
    Finance,
    ComplianceShare,
)

from common.permission import (
    IsLandlord,
    IsMortgageAdviser,
    IsAdmin,
    CanAccessProperty,
    CanAccessMortgage,
)

from api.serializers.property import (
    PropertySerializer,
    MortgageSerializers,
    TenantSerializer,
    ComplianceAndCertificationSerializers,
    UploadDocumentSerializer,
    FinanceSerializer,
    PropertyOnboardingSerializer,
    ComplianceShareSerializer,
)


class PropertyListView(ListCreateAPIView):
    serializer_class = PropertySerializer
    permission_classes = [CanAccessProperty]
    filterset_fields = ["property_type", "status"]
    search_fields = ["property_name", "address"]

    def get_queryset(self):
        organisation = self.request.user.get_organisation()

        if not organisation:
            raise NotFound("Organisation not found for the user.")

        queryset = Property.objects.filter(organisation=organisation)

        current_user = OrganisationUser.objects.filter(
            user=self.request.user,
            organisation=organisation,
        ).first()

        if current_user.role in [
            OrganisationRoleChoices.LANDLORD,
            OrganisationRoleChoices.ADMIN,
        ]:
            return queryset
        else:
            queryset = queryset.filter(
                property_permissions__user=self.request.user,
                property_permissions__can_view=True,
            ).distinct()

        return queryset

    def perform_create(self, serializer):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        serializer.save(organisation=organisation)


class PropertyDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = PropertySerializer
    permission_classes = [CanAccessProperty]

    def get_object(self):
        obj = get_object_or_404(
            Property,
            alias=self.kwargs["property_alias"],
        )

        # Check if the user has permission to access this property
        self.check_object_permissions(self.request, obj)

        return obj


class MortgageListView(ListCreateAPIView):
    serializer_class = MortgageSerializers
    permission_classes = [CanAccessMortgage]
    search_fields = ["property__property_name", "lender_name"]

    def get_queryset(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")

        queryset = Mortgage.objects.filter(organisation=organisation)

        # Mortgage advisers can only see permitted properties
        current_user = OrganisationUser.objects.filter(
            user=self.request.user,
            organisation=organisation,
        ).first()

        if current_user.role in [
            OrganisationRoleChoices.LANDLORD,
            OrganisationRoleChoices.ADMIN,
        ]:
            return queryset
        else:
            queryset = queryset.filter(
                mortgage_permissions__user=self.request.user,
                mortgage_permissions__can_view=True,
            ).distinct()

    def perform_create(self, serializer):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        serializer.save(organisation=organisation)


class MortgageDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = MortgageSerializers
    permission_classes = [CanAccessMortgage]

    def get_object(self):
        obj = get_object_or_404(Mortgage, alias=self.kwargs["mortgage_alias"])

        # Check if the user has permission to access this mortgage
        self.check_object_permissions(self.request, obj)

        return obj


class TenantListView(ListCreateAPIView):
    serializer_class = TenantSerializer
    permission_classes = [IsLandlord | IsAdmin]
    search_fields = [
        "property__property_name",
        "first_name",
        "last_name",
        "email",
        "phone",
    ]

    def get_queryset(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        return Tenant.objects.filter(organisation=organisation).order_by("-created_at")

    def perform_create(self, serializer):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        serializer.save(organisation=organisation, password=make_password(None))


class TenantDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = TenantSerializer
    permission_classes = [IsLandlord | IsAdmin]

    def get_object(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        return get_object_or_404(
            Tenant, alias=self.kwargs["tenant_alias"], organisation=organisation
        )


class ComplianceAndCertificationListView(ListCreateAPIView):
    serializer_class = ComplianceAndCertificationSerializers
    permission_classes = [IsLandlord | IsAdmin]
    search_fields = ["property__property_name", "certificate_number"]

    def get_queryset(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        return ComplianceAndCertification.objects.filter(
            organisation=organisation
        ).order_by("-created_at")

    def perform_create(self, serializer):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        serializer.save(organisation=organisation)


class ComplianceAndCertificationDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = ComplianceAndCertificationSerializers
    permission_classes = [IsLandlord | IsAdmin]

    def get_object(self):
        return get_object_or_404(
            ComplianceAndCertification, alias=self.kwargs["compliance_alias"]
        )


class UploadDocumentListCreateApiView(ListCreateAPIView):
    serializer_class = UploadDocumentSerializer
    permission_classes = [IsLandlord | IsAdmin]
    search_fields = ["property__property_name", "document_name"]
    filterset_fields = ["document_category"]

    def get_queryset(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        return UploadDocument.objects.filter(organisation=organisation)

    def perform_create(self, serializer):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")

        uploaded_files = self.request.FILES.getlist("uploaded_files")
        serializer.save(organisation=organisation, uploaded_files=uploaded_files)


class UploadDocumentRetrieveAPIView(RetrieveUpdateDestroyAPIView):
    serializer_class = UploadDocumentSerializer
    permission_classes = [IsLandlord | IsAdmin]

    def get_object(self):
        return get_object_or_404(UploadDocument, alias=self.kwargs["document_alias"])

    def perform_update(self, serializer):
        uploaded_files = self.request.FILES.getlist("uploaded_files")
        serializer.save(uploaded_files=uploaded_files)


class FinanceListView(ListCreateAPIView):
    serializer_class = FinanceSerializer
    permission_classes = [IsLandlord | IsAdmin]
    search_fields = ["property__property_name"]

    def get_queryset(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        return Finance.objects.filter(organisation=organisation)

    def perform_create(self, serializer):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        serializer.save(organisation=organisation)


class FinanceDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = FinanceSerializer
    permission_classes = [IsLandlord | IsAdmin]

    def get_object(self):
        return get_object_or_404(Finance, alias=self.kwargs["finance_alias"])


class PropertyOnboardingAPIView(APIView):
    permission_classes = [IsLandlord | IsAdmin]

    def post(self, request, *args, **kwargs):
        serializer = PropertyOnboardingSerializer(
            data=request.data,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)
        return Response(serializer.save(), status=status.HTTP_200_OK)


class PropertyPortfolioExportView(APIView):
    permission_classes = [IsLandlord | IsAdmin | IsMortgageAdviser]

    HEADERS = [
        "Property Address",
        "Property Type",
        "Number Of Bedrooms",
        "Owners",
        "Original Purchase Date",
        "Original Purchase Price",
        "Current Market Value",
        "Mortgage Lender Name",
        "Outstanding Balance",
        "Current Interest Rate",
        "Monthly Mortgage Payment",
        "Repayment Method",
        "Mortgage End Date",
        "Interest Rate Expiry Date",
        "Monthly Rental Income",
        "EPC Rating",
        "Property Tenure",
        "If Leasehold, Remaining Lease Term",
        "Monthly Service Charge",
    ]

    def get(self, request):
        export_format = request.query_params.get("export_format", "xlsx").lower()
        if export_format not in ("xlsx", "pdf"):
            return HttpResponse(
                "Invalid format. Use ?export_format=xlsx or ?export_format=pdf.",
                status=400,
            )

        organisation = request.user.get_organisation()
        if not organisation:
            return HttpResponse("You are not part of any organisation.", status=400)

        properties = (
            Property.objects.filter(organisation=organisation)
            .prefetch_related("property_mortgages", "shareholder")
            .order_by("-created_at")
        )

        rows = self._build_rows(properties)
        stamp = date.today().isoformat()

        if export_format == "xlsx":
            buffer = self._build_xlsx(rows)
            filename = f"Property_Portfolio_Summary_-_Landkeeper_{stamp}.xlsx"
            content_type = (
                "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
            )
        else:
            buffer = self._build_pdf(rows)
            filename = f"Property_Portfolio_Summary_-_Landkeeper_{stamp}.pdf"
            content_type = "application/pdf"

        response = HttpResponse(buffer.getvalue(), content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # Data assembly
    def _build_rows(self, properties):
        rows = []
        for prop in properties:
            mortgage = prop.property_mortgages.order_by("-created_at").first()
            owners = ", ".join(
                filter(None, (o.owner_name for o in prop.shareholder.all()))
            )

            rows.append(
                {
                    "address": prop.address,
                    "property_type": prop.get_property_type_display(),
                    "bedrooms": prop.bedrooms,
                    "owners": owners,
                    "purchase_date": prop.purchase_date,
                    "purchase_price": prop.purchase_price,
                    "current_value": prop.current_value,
                    "lender": mortgage.lender_name if mortgage else "",
                    "outstanding_balance": (
                        mortgage.outstanding_balance if mortgage else None
                    ),
                    "interest_rate": mortgage.interest_rate if mortgage else None,
                    "monthly_payment": mortgage.monthly_payment if mortgage else None,
                    "repayment_method": (
                        mortgage.get_interest_rate_type_display() if mortgage else ""
                    ),
                    "mortgage_end_date": None,
                    "rate_expiry_date": (
                        mortgage.interest_rate_expiry_date if mortgage else None
                    ),
                    "monthly_rental_income": prop.monthly_rental_income,
                    "epc": mortgage.epc_rating if mortgage else "",
                    "tenure": (
                        prop.get_property_tenure_display()
                        if prop.property_tenure
                        else ""
                    ),
                    "lease_term": prop.remaining_lease_term,
                    "service_charge": prop.monthly_service_charge,
                }
            )
        return rows

    def _totals(self, rows):
        def total(key):
            return sum((r[key] or 0) for r in rows)

        current_value = total("current_value")
        outstanding_balance = total("outstanding_balance")
        monthly_payment = total("monthly_payment")
        monthly_rental_income = total("monthly_rental_income")
        return {
            "current_value": current_value,
            "outstanding_balance": outstanding_balance,
            "monthly_payment": monthly_payment,
            "monthly_rental_income": monthly_rental_income,
            "net_asset_value": current_value - outstanding_balance,
            "net_income": monthly_rental_income - monthly_payment,
        }

    # Excel export (matches template)
    def _build_xlsx(self, rows):
        LIGHT_BLUE = "DCE6F1"
        GREY = "D9D9D9"

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        header_font = Font(name="Calibri", size=10, bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        body_font = Font(name="Calibri", size=10)
        thin = Side(style="thin", color=GREY)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, label in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[1].height = 42

        widths = [
            30,
            14,
            10,
            18,
            14,
            15,
            14,
            18,
            15,
            14,
            16,
            16,
            14,
            16,
            15,
            10,
            14,
            18,
            14,
        ]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        currency_cols = {6, 7, 9, 11, 15, 19}
        percent_cols = {10}
        date_cols = {5, 13, 14}

        row_idx = 2
        for r in rows:
            values = [
                r["address"],
                r["property_type"],
                r["bedrooms"],
                r["owners"],
                r["purchase_date"],
                r["purchase_price"],
                r["current_value"],
                r["lender"],
                r["outstanding_balance"],
                r["interest_rate"],
                r["monthly_payment"],
                r["repayment_method"],
                r["mortgage_end_date"],
                r["rate_expiry_date"],
                r["monthly_rental_income"],
                r["epc"],
                r["tenure"],
                r["lease_term"],
                r["service_charge"],
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if col == 1 else "center", vertical="center"
                )
                if col in currency_cols:
                    cell.number_format = "£#,##0.00;(£#,##0.00);-"
                elif col in percent_cols:
                    cell.number_format = "0.00%"
                elif col in date_cols:
                    cell.number_format = "dd/mm/yyyy"
            row_idx += 1

        last_data_row = row_idx - 1

        # TOTALS label row (with sub-labels: VALUES / MORTGAGE SUM / PAYMENTS / RENT)
        label_row = row_idx
        ws.cell(row=label_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=label_row, column=7, value="VALUES").font = Font(bold=True)
        ws.cell(row=label_row, column=9, value="MORTGAGE SUM").font = Font(bold=True)
        ws.cell(row=label_row, column=11, value="PAYMENTS").font = Font(bold=True)
        ws.cell(row=label_row, column=15, value="RENT").font = Font(bold=True)
        for col in range(1, len(self.HEADERS) + 1):
            cell = ws.cell(row=label_row, column=col)
            cell.fill = PatternFill("solid", fgColor=GREY)
            cell.border = border

        # Totals values row
        values_row = label_row + 1
        ws.cell(row=values_row, column=7, value=f"=SUM(G2:G{last_data_row})")
        ws.cell(row=values_row, column=9, value=f"=SUM(I2:I{last_data_row})")
        ws.cell(row=values_row, column=11, value=f"=SUM(K2:K{last_data_row})")
        ws.cell(row=values_row, column=15, value=f"=SUM(O2:O{last_data_row})")
        for col in (7, 9, 11, 15):
            cell = ws.cell(row=values_row, column=col)
            cell.font = Font(bold=True)
            cell.number_format = "£#,##0.00"
            cell.fill = PatternFill("solid", fgColor=GREY)
            cell.border = border

        # NET ASSET VALUE / NET INCOME rows
        nav_row = values_row + 1
        income_row = values_row + 2
        ws.cell(row=nav_row, column=1, value="NET ASSET VALUE")
        ws.cell(row=nav_row, column=2, value=f"=G{values_row}-I{values_row}")
        ws.cell(row=income_row, column=1, value="NET INCOME")
        ws.cell(row=income_row, column=2, value=f"=O{values_row}-K{values_row}")
        for r in (nav_row, income_row):
            for col in (1, 2):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                if col == 2:
                    cell.number_format = "£#,##0.00"

        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # PDF export
    def _build_pdf(self, rows):
        INK = colors.HexColor("#111827")
        SUBTLE = colors.HexColor("#6B7280")
        HAIRLINE = colors.HexColor("#E5E7EB")
        ACCENT = colors.HexColor("#8A6D3B")
        WHITE = colors.white
        ZEBRA = colors.HexColor("#F9FAFB")

        page_size = landscape(A4)
        page_w, _ = page_size

        title_style = ParagraphStyle(
            "Title",
            fontName="Helvetica-Bold",
            fontSize=20,
            textColor=INK,
            alignment=TA_LEFT,
            leading=24,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            fontName="Helvetica",
            fontSize=10,
            textColor=SUBTLE,
        )
        kpi_label_style = ParagraphStyle(
            "KPILabel",
            fontName="Helvetica",
            fontSize=9,
            textColor=SUBTLE,
        )
        kpi_value_style = ParagraphStyle(
            "KPIValue",
            fontName="Helvetica-Bold",
            fontSize=17,
            textColor=INK,
            spaceBefore=2,
        )
        header_left = ParagraphStyle(
            "HeaderLeft",
            fontName="Helvetica-Bold",
            fontSize=6.6,
            textColor=INK,
            leading=8,
        )
        header_center = ParagraphStyle(
            "HeaderCenter",
            fontName="Helvetica-Bold",
            fontSize=6.6,
            textColor=INK,
            alignment=1,
            leading=8,
        )
        # Wrapping body styles so long text doesn't get clipped/overflow the cell
        body_left = ParagraphStyle(
            "BodyLeft",
            fontName="Helvetica",
            fontSize=7.6,
            textColor=INK,
            alignment=TA_LEFT,
            leading=9,
        )
        body_center = ParagraphStyle(
            "BodyCenter",
            fontName="Helvetica",
            fontSize=7.6,
            textColor=INK,
            alignment=1,
            leading=9,
        )

        def fmt_money(v):
            return f"£{v:,.0f}" if v is not None else ""

        def fmt_pct(v):
            return f"{v:.2f}%" if v is not None else ""

        def fmt_date(v):
            return v.strftime("%d/%m/%Y") if v else ""

        # Columns whose text can run long and need wrapping rather than truncation
        wrap_left_cols = {
            0,
            3,
            7,
            11,
            16,
        }  # address, owners, lender, repayment method, tenure

        table_data = [
            [
                Paragraph(label, header_left if i == 0 else header_center)
                for i, label in enumerate(self.HEADERS)
            ]
        ]
        for r in rows:
            raw_values = [
                r["address"] or "",
                r["property_type"] or "",
                str(r["bedrooms"]) if r["bedrooms"] is not None else "",
                r["owners"] or "",
                fmt_date(r["purchase_date"]),
                fmt_money(r["purchase_price"]),
                fmt_money(r["current_value"]),
                r["lender"] or "",
                fmt_money(r["outstanding_balance"]),
                fmt_pct(r["interest_rate"]),
                fmt_money(r["monthly_payment"]),
                r["repayment_method"] or "",
                fmt_date(r["mortgage_end_date"]),
                fmt_date(r["rate_expiry_date"]),
                fmt_money(r["monthly_rental_income"]),
                r["epc"] or "",
                r["tenure"] or "",
                str(r["lease_term"]) if r["lease_term"] is not None else "",
                fmt_money(r["service_charge"]),
            ]
            row_cells = []
            for i, value in enumerate(raw_values):
                if i in wrap_left_cols:
                    row_cells.append(Paragraph(value, body_left))
                else:
                    row_cells.append(Paragraph(value, body_center))
            table_data.append(row_cells)

        totals = self._totals(rows)
        totals_row = [
            "Totals",
            "",
            "",
            "",
            "",
            "",
            fmt_money(totals["current_value"]),
            "",
            fmt_money(totals["outstanding_balance"]),
            "",
            fmt_money(totals["monthly_payment"]),
            "",
            "",
            "",
            fmt_money(totals["monthly_rental_income"]),
            "",
            "",
            "",
            "",
        ]
        table_data.append(totals_row)

        col_widths = [
            33 * mm,
            17 * mm,
            12 * mm,
            14 * mm,
            16 * mm,
            16 * mm,
            17 * mm,
            16 * mm,
            16 * mm,
            13 * mm,
            15 * mm,
            18 * mm,
            15 * mm,
            15 * mm,
            17 * mm,
            11 * mm,
            14 * mm,
            13 * mm,
            15 * mm,
        ]
        scale = (page_w - 20 * mm) / sum(col_widths)
        col_widths = [w * scale for w in col_widths]

        n_body_rows = len(rows)
        table_style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), WHITE),
            ("VALIGN", (0, 0), (-1, 0), "BOTTOM"),
            ("TOPPADDING", (0, 0), (-1, 0), 4),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("LEFTPADDING", (0, 0), (-1, 0), 3),
            ("RIGHTPADDING", (0, 0), (-1, 0), 3),
            ("LINEBELOW", (0, 0), (-1, 0), 1, INK),
            ("VALIGN", (0, 1), (-1, -2), "MIDDLE"),
            ("TOPPADDING", (0, 1), (-1, -2), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -2), 6),
            ("LEFTPADDING", (0, 1), (-1, -2), 3),
            ("RIGHTPADDING", (0, 1), (-1, -2), 3),
            ("LINEBELOW", (0, 1), (-1, -2), 0.5, HAIRLINE),
            ("LINEABOVE", (0, -1), (-1, -1), 1, INK),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, -1), (-1, -1), 8),
            ("ALIGN", (0, -1), (-1, -1), "CENTER"),
            ("ALIGN", (0, -1), (0, -1), "LEFT"),
            ("TOPPADDING", (0, -1), (-1, -1), 7),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 7),
        ]
        # Zebra striping on body rows for readability
        for i in range(1, n_body_rows + 1):
            if i % 2 == 0:
                table_style_cmds.append(("BACKGROUND", (0, i), (-1, i), ZEBRA))

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_style_cmds))

        def header_footer(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(HAIRLINE)
            canvas.setLineWidth(0.5)
            canvas.line(10 * mm, 12 * mm, page_w - 10 * mm, 12 * mm)
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(SUBTLE)
            canvas.drawString(
                10 * mm, 8 * mm, "Landkeeper  ·  Property Portfolio Summary"
            )
            canvas.drawRightString(page_w - 10 * mm, 8 * mm, f"Page {doc.page}")
            canvas.restoreState()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=14 * mm,
            bottomMargin=16 * mm,
            title="Property Portfolio Summary - Landkeeper",
        )

        story = [
            Paragraph("Property Portfolio Summary", title_style),
            Spacer(1, 3),
            Paragraph("Landkeeper — Landlord Property Management", subtitle_style),
            Spacer(1, 14),
            table,
            Spacer(1, 16),
        ]

        kpi_table = Table(
            [
                [
                    Paragraph("NET ASSET VALUE", kpi_label_style),
                    Paragraph("NET INCOME", kpi_label_style),
                ],
                [
                    Paragraph(fmt_money(totals["net_asset_value"]), kpi_value_style),
                    Paragraph(fmt_money(totals["net_income"]), kpi_value_style),
                ],
            ],
            colWidths=[70 * mm, 70 * mm],
        )
        kpi_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LINEABOVE", (0, 0), (0, 0), 1.2, ACCENT),
                    ("LINEABOVE", (1, 0), (1, 0), 1.2, ACCENT),
                    ("TOPPADDING", (0, 0), (-1, 0), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 0),
                    ("TOPPADDING", (0, 1), (-1, 1), 2),
                ]
            )
        )
        story.append(kpi_table)

        doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
        buffer.seek(0)
        return buffer


class ComplianceAndCertificationShareView(APIView):
    serializer_class = ComplianceShareSerializer
    permission_classes = [IsLandlord | IsAdmin]

    def get_compliance(self):
        organisation = self.request.user.get_organisation()
        if not organisation:
            raise NotFound("Organisation not found for the user.")
        return get_object_or_404(
            ComplianceAndCertification,
            alias=self.kwargs["compliance_alias"],
            organisation=organisation,
        )

    def _resolve_tenants(self, compliance, tenant_aliases):
        tenants = Tenant.objects.filter(
            alias__in=tenant_aliases,
            organisation=compliance.organisation,
        )
        found_aliases = {str(a) for a in tenants.values_list("alias", flat=True)}
        missing = [a for a in tenant_aliases if str(a) not in found_aliases]
        if missing:
            raise ValidationError(
                {"tenant": f"Unknown tenant alias(es): {', '.join(missing)}"}
            )
        return tenants

    def post(self, request, *args, **kwargs):
        compliance = self.get_compliance()
        serializer = ComplianceShareSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_aliases = serializer.validated_data["tenant"]

        tenants = self._resolve_tenants(compliance, tenant_aliases)

        for tenant in tenants:
            ComplianceShare.objects.get_or_create(
                compliance=compliance,
                tenant=tenant,
            )

        return Response(
            {"detail": "Compliance certificate shared."},
            status=status.HTTP_200_OK,
        )

    def delete(self, request, *args, **kwargs):
        compliance = self.get_compliance()
        serializer = ComplianceShareSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_aliases = serializer.validated_data["tenant"]

        tenants = self._resolve_tenants(compliance, tenant_aliases)
        deleted_count, _ = ComplianceShare.objects.filter(
            compliance=compliance, tenant__in=tenants
        ).delete()

        return Response(
            {"detail": "Access revoked.", "revoked_count": deleted_count},
            status=status.HTTP_200_OK,
        )